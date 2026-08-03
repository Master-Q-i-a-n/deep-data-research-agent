#!/usr/bin/env python3
"""Deterministically profile CSV, TSV and XLSX inputs without business analysis."""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import math
import re
from collections.abc import Iterable, Iterator, Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

MAX_PROFILE_ROWS = 100_000
MAX_SAMPLE_ROWS = 5
MAX_UNIQUE_TRACKED = 10_000
SUPPORTED_SUFFIXES = {".csv", ".tsv", ".xlsx"}
SENSITIVE_COLUMN = re.compile(
    r"(password|token|secret|id.?card|phone|mobile|email|address|name|"
    r"密码|令牌|密钥|身份证|手机|电话|邮箱|地址|姓名)",
    re.IGNORECASE,
)
EMAIL_VALUE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
PHONE_VALUE = re.compile(r"^\+?\d[\d\s-]{7,}\d$")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, str)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def _is_missing(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _infer_value_type(value: Any) -> str:
    if _is_missing(value):
        return "missing"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (datetime, date)):
        return "datetime"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    text = str(value).strip()
    # Numeric-looking identifiers with leading zero stay strings.
    if re.fullmatch(r"[+-]?\d+", text) and not re.fullmatch(r"[+-]?0\d+", text):
        return "integer"
    if re.fullmatch(r"[+-]?(?:\d+\.\d*|\d*\.\d+)", text):
        return "number"
    try:
        datetime.fromisoformat(text)
    except ValueError:
        return "string"
    return "datetime"


def _safe_field_names(header: Sequence[Any], width: int) -> list[str]:
    names: list[str] = []
    used: dict[str, int] = {}
    for index in range(width):
        raw = (
            str(header[index]).strip()
            if index < len(header) and header[index] is not None
            else ""
        )
        base = raw or f"column_{index + 1}"
        count = used.get(base, 0) + 1
        used[base] = count
        names.append(base if count == 1 else f"{base}_{count}")
    return names


def _redacted_sample(field: str, value: Any) -> Any:
    normalized = _json_value(value)
    text = str(normalized or "").strip()
    if (
        SENSITIVE_COLUMN.search(field)
        or EMAIL_VALUE.fullmatch(text)
        or PHONE_VALUE.fullmatch(text)
    ):
        return "<已脱敏>"
    if len(text) > 120:
        return text[:117] + "..."
    return normalized


def _profile_rows(
    rows: Iterable[Sequence[Any]],
    *,
    estimated_rows: int | None = None,
    estimated_columns: int | None = None,
) -> dict[str, Any]:
    iterator = iter(rows)
    header: list[Any] | None = None
    header_row = 0
    for row_number, row in enumerate(iterator, start=1):
        values = list(row)
        if any(not _is_missing(value) for value in values):
            header = values
            header_row = row_number
            break
    if header is None:
        return {
            "blank": True,
            "estimated_rows": estimated_rows,
            "estimated_columns": estimated_columns or 0,
            "scanned_rows": 0,
            "truncated": False,
            "candidate_header": None,
            "fields": [],
            "duplicate_rows": 0,
            "sample_rows": [],
        }

    width = max(len(header), estimated_columns or 0)
    buffered: list[list[Any]] = []
    truncated = False
    for row in iterator:
        if len(buffered) >= MAX_PROFILE_ROWS:
            truncated = True
            break
        values = list(row)
        width = max(width, len(values))
        buffered.append(values)

    fields = _safe_field_names(header, width)
    stats = [
        {
            "name": field,
            "position": index + 1,
            "missing": 0,
            "non_null": 0,
            "types": {},
            "unique": set(),
            "unique_truncated": False,
            "numeric_min": None,
            "numeric_max": None,
            "date_min": None,
            "date_max": None,
        }
        for index, field in enumerate(fields)
    ]
    row_hashes: set[bytes] = set()
    duplicates = 0
    samples: list[dict[str, Any]] = []

    for row in buffered:
        normalized_row = [
            row[index] if index < len(row) else None for index in range(width)
        ]
        fingerprint = hashlib.sha256(
            json.dumps(
                [_json_value(value) for value in normalized_row], ensure_ascii=False
            ).encode("utf-8")
        ).digest()
        if fingerprint in row_hashes:
            duplicates += 1
        else:
            row_hashes.add(fingerprint)

        if len(samples) < MAX_SAMPLE_ROWS:
            samples.append(
                {
                    field: _redacted_sample(field, normalized_row[index])
                    for index, field in enumerate(fields)
                }
            )

        for index, value in enumerate(normalized_row):
            stat = stats[index]
            value_type = _infer_value_type(value)
            stat["types"][value_type] = stat["types"].get(value_type, 0) + 1
            if value_type == "missing":
                stat["missing"] += 1
                continue
            stat["non_null"] += 1
            serialized = json.dumps(_json_value(value), ensure_ascii=False)
            if len(stat["unique"]) < MAX_UNIQUE_TRACKED:
                stat["unique"].add(serialized)
            elif serialized not in stat["unique"]:
                stat["unique_truncated"] = True

            if value_type in {"integer", "number"}:
                number = float(str(value).strip())
                stat["numeric_min"] = (
                    number
                    if stat["numeric_min"] is None
                    else min(stat["numeric_min"], number)
                )
                stat["numeric_max"] = (
                    number
                    if stat["numeric_max"] is None
                    else max(stat["numeric_max"], number)
                )
            elif value_type == "datetime":
                rendered = _json_value(value)
                stat["date_min"] = (
                    rendered
                    if stat["date_min"] is None
                    else min(stat["date_min"], rendered)
                )
                stat["date_max"] = (
                    rendered
                    if stat["date_max"] is None
                    else max(stat["date_max"], rendered)
                )

    field_profiles: list[dict[str, Any]] = []
    for stat in stats:
        types = dict(stat.pop("types"))
        unique = stat.pop("unique")
        non_missing_types = {
            name: count for name, count in types.items() if name != "missing"
        }
        inferred = (
            max(non_missing_types, key=non_missing_types.get)
            if non_missing_types
            else "unknown"
        )
        field_profiles.append(
            {
                **stat,
                "inferred_type": inferred,
                "type_counts": types,
                "unique_values_scanned": len(unique),
            }
        )

    if estimated_rows is not None and estimated_rows > len(buffered) + header_row:
        truncated = truncated or len(buffered) >= MAX_PROFILE_ROWS
    return {
        "blank": False,
        "estimated_rows": estimated_rows,
        "estimated_columns": estimated_columns or width,
        "scanned_rows": len(buffered),
        "truncated": truncated,
        "candidate_header": {
            "row": header_row,
            "values": [_json_value(value) for value in header],
        },
        "fields": field_profiles,
        "duplicate_rows": duplicates,
        "sample_rows": samples,
    }


def _decode_delimited(path: Path) -> tuple[str, str]:
    content = path.read_bytes()
    if content.startswith(b"\xef\xbb\xbf"):
        return content.decode("utf-8-sig"), "utf-8-sig"
    for encoding in ("utf-8", "gb18030"):
        try:
            return content.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise ValueError("无法使用 UTF-8、UTF-8-SIG 或 GB18030 解码文件")


def _profile_delimited(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    text, encoding = _decode_delimited(path)
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
    table = _profile_rows(rows)
    table.update({"name": path.stem, "state": "visible", "encoding": encoding})
    return [table], []


def _worksheet_rows(
    formula_sheet: Any,
    cached_sheet: Any,
    counters: dict[str, int],
) -> Iterator[list[Any]]:
    cached_rows = cached_sheet.iter_rows(values_only=True)
    for formula_row in formula_sheet.iter_rows():
        cached_row = next(cached_rows, ())
        values: list[Any] = []
        for index, cell in enumerate(formula_row):
            cached = cached_row[index] if index < len(cached_row) else None
            if cell.data_type == "f":
                counters["formula_cells"] += 1
                if cached is None:
                    counters["formula_cells_without_cache"] += 1
            values.append(cached if cell.data_type == "f" else cell.value)
        yield values


def _profile_xlsx(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    formulas = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    cached = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    tables: list[dict[str, Any]] = []
    warnings: list[str] = []
    try:
        for formula_sheet in formulas.worksheets:
            cached_sheet = cached[formula_sheet.title]
            counters = {
                "formula_cells": 0,
                "formula_cells_without_cache": 0,
            }

            table = _profile_rows(
                _worksheet_rows(formula_sheet, cached_sheet, counters),
                estimated_rows=formula_sheet.max_row,
                estimated_columns=formula_sheet.max_column,
            )
            table.update(
                {
                    "name": formula_sheet.title,
                    "state": formula_sheet.sheet_state,
                    **counters,
                }
            )
            if counters["formula_cells_without_cache"]:
                warnings.append(
                    f"工作表 {formula_sheet.title} 有 "
                    f"{counters['formula_cells_without_cache']} 个公式缺少缓存值"
                )
            tables.append(table)
    finally:
        formulas.close()
        cached.close()
    return tables, warnings


def build_profile(path: Path) -> dict[str, Any]:
    path = path.resolve()
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError("仅支持 CSV、TSV 和 XLSX 文件")
    if not path.is_file():
        raise FileNotFoundError(f"输入文件不存在：{path}")

    tables, warnings = (
        _profile_xlsx(path) if suffix == ".xlsx" else _profile_delimited(path)
    )
    if any(table["truncated"] for table in tables):
        warnings.append("至少一张表的探查达到 100,000 行上限；正式分析需处理完整数据")
    return {
        "status": "success",
        "input": str(path),
        "sha256": _sha256(path),
        "format": suffix.removeprefix("."),
        "size": path.stat().st_size,
        "tables": tables,
        "warnings": warnings,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="探查 CSV、TSV 或 XLSX 表格结构")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    output = Path(args.output)
    try:
        profile = build_profile(Path(args.input))
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(
            json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(
            json.dumps(
                {
                    "status": "success",
                    "profile": str(output),
                    "tables": len(profile["tables"]),
                    "warning_count": len(profile["warnings"]),
                },
                ensure_ascii=False,
            )
        )
        return 0
    except Exception as exc:  # noqa: BLE001 - CLI reports a compact deterministic error.
        print(json.dumps({"status": "failed", "error": str(exc)}, ensure_ascii=False))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
